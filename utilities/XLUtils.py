import openpyxl
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)
def readData(file,sheetName,rownum,columnnu):
    workbook  = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=columnnu).value

def writeData(file,sheeName,rownum,columnnu,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheeName]
    sheet.cell(row=rownum,column=columnnu).value = data
    workbook.save(file)


